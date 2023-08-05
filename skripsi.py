from flask import Flask, jsonify, request
from werkzeug.utils import secure_filename
from flask_cors import CORS
from PIL import Image
from resnet import detect
import numpy as np
import cv2
import os
import json
import datetime
import openpyxl

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "http://localhost:3000"}})

@app.route('/detect', methods=['POST'])
def detect_deep_fake():
    model = 'model.pth' 

    file = request.files['files']
    filename = secure_filename(file.filename)
    file.save(filename)
    extension = os.path.splitext(filename)[1]

    if extension == ".mp4" or extension == ".avi" or extension == ".mov" or extension == ".wmv":
        cap = cv2.VideoCapture(filename)
        if not cap.isOpened():
            return jsonify("Error opening video file")

        frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        fps = cap.get(cv2.CAP_PROP_FPS)
        frame_skip = int(fps) * 3
        predictions = []

        for i in range(frame_count):
            ret, frame = cap.read()
            if ret:
                if i % frame_skip == 0: 
                    net = detect('cpu', model)
                    img = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
                    img.load()
                    logit = net.apply(img)

                    prediction = logit.tolist()
                    predictions.append(float(prediction))

        final_result = np.mean(predictions, axis=0)
        cap.release()

    else:
        net = detect('cpu', model)
        img = Image.open(file).convert('RGB')
        img.load()
        logit = net.apply(img)

        final_result = logit.tolist()

    if (final_result > 0):
        result = {'prediction': 'fake', 'percentage': float(( final_result + 50 ) / 100)}
    else:
        result = {'prediction': 'real', 'percentage': float(( 50 - final_result ) / 100)}

    os.remove(filename)
    return jsonify(result)


@app.route('/message', methods=['POST'])
def receive_message():
    json_data = request.json
    time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    content = json_data['message']

    try:
        workbook = openpyxl.load_workbook('message.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active
    next_row = sheet.max_row + 1

    sheet.cell(row=next_row, column=1).value = time
    sheet.cell(row=next_row, column=2).value = content

    workbook.save('message.xlsx')
    return 'Message received and saved'


if (__name__ == '__main__'):
    app.run()